"""
Complete pipeline endpoints for domain verification.
Includes Stage 1 (API), Stage 2 (Playwright), Stage 3 (TXT setup), and Stage 4 (TXT execution).
Enhanced with CSV upload, WebSocket progress updates, and external API integration.
"""
import asyncio
import os
import csv
import json
from datetime import datetime, timezone
from typing import List, Optional, Dict, Any
from pathlib import Path
from fastapi import APIRouter, HTTPException, BackgroundTasks, UploadFile, File, WebSocket, WebSocketDisconnect
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel
import httpx

import sys
sys.path.insert(0, str(Path(__file__).parent.parent.parent.parent))

from complete_domain_pipeline import CompleteDomainPipeline
from config.settings import settings

router = APIRouter(prefix="/api/pipeline", tags=["pipeline"])

# WebSocket connection manager
class ConnectionManager:
    def __init__(self):
        self.active_connections: Dict[str, WebSocket] = {}

    async def connect(self, run_id: str, websocket: WebSocket):
        await websocket.accept()
        self.active_connections[run_id] = websocket

    def disconnect(self, run_id: str):
        if run_id in self.active_connections:
            del self.active_connections[run_id]

    async def send_progress(self, run_id: str, message: dict):
        if run_id in self.active_connections:
            try:
                await self.active_connections[run_id].send_json(message)
            except:
                self.disconnect(run_id)

manager = ConnectionManager()


# Request/Response Models
class PipelineRunRequest(BaseModel):
    """Request model for running complete pipeline."""
    domains: List[str]
    enable_txt_verification: bool = False
    txt_wait_time: int = 30
    txt_max_attempts: int = 1
    txt_poll_interval: int = 30


class PipelineRunResponse(BaseModel):
    """Response model for pipeline run."""
    run_id: str
    status: str
    message: str
    domains_count: int
    csv_download_url: Optional[str] = None
    report_url: Optional[str] = None
    websocket_url: Optional[str] = None


class PipelineStatusResponse(BaseModel):
    """Response model for pipeline status."""
    run_id: str
    status: str
    stage: Optional[str] = None
    progress: Optional[dict] = None
    results_available: bool = False
    csv_url: Optional[str] = None
    report_url: Optional[str] = None
    error: Optional[str] = None


class ExternalAPIConfig(BaseModel):
    """Configuration for external API endpoints."""
    momen_api_url: Optional[str] = None
    frontend_api_url: Optional[str] = None
    momen_api_key: Optional[str] = None
    frontend_api_key: Optional[str] = None


# In-memory storage for pipeline status (use Redis/DB in production)
pipeline_status = {}

# External API configuration (can be set via environment variables)
external_api_config = ExternalAPIConfig(
    momen_api_url=os.getenv("MOMEN_API_URL"),
    frontend_api_url=os.getenv("FRONTEND_API_URL"),
    momen_api_key=os.getenv("MOMEN_API_KEY"),
    frontend_api_key=os.getenv("FRONTEND_API_KEY")
)


async def send_to_external_apis(run_id: str, csv_path: Path):
    """Send results to external APIs (momen and frontend)."""
    try:
        # Read CSV content
        with open(csv_path, 'r', encoding='utf-8') as f:
            csv_content = f.read()
        
        # Parse CSV to JSON
        csv_reader = csv.DictReader(csv_content.splitlines())
        results_json = list(csv_reader)
        
        # Send to Momen API (if configured)
        if external_api_config.momen_api_url:
            try:
                async with httpx.AsyncClient(timeout=30.0) as client:
                    headers = {}
                    if external_api_config.momen_api_key:
                        headers["Authorization"] = f"Bearer {external_api_config.momen_api_key}"
                    
                    response = await client.post(
                        external_api_config.momen_api_url,
                        json={
                            "run_id": run_id,
                            "results": results_json,
                            "timestamp": datetime.now(timezone.utc).isoformat()
                        },
                        headers=headers
                    )
                    print(f"✅ Sent results to Momen API: {response.status_code}")
            except Exception as e:
                print(f"⚠️  Failed to send to Momen API: {str(e)}")
        
        # Send to Frontend API (if configured)
        if external_api_config.frontend_api_url:
            try:
                async with httpx.AsyncClient(timeout=30.0) as client:
                    headers = {}
                    if external_api_config.frontend_api_key:
                        headers["Authorization"] = f"Bearer {external_api_config.frontend_api_key}"
                    
                    response = await client.post(
                        external_api_config.frontend_api_url,
                        json={
                            "run_id": run_id,
                            "results": results_json,
                            "csv_url": f"/api/pipeline/{run_id}/csv",
                            "timestamp": datetime.now(timezone.utc).isoformat()
                        },
                        headers=headers
                    )
                    print(f"✅ Sent results to Frontend API: {response.status_code}")
            except Exception as e:
                print(f"⚠️  Failed to send to Frontend API: {str(e)}")
                
    except Exception as e:
        print(f"❌ Error sending to external APIs: {str(e)}")


async def run_pipeline_task(
    run_id: str,
    domains: List[str],
    enable_txt: bool,
    txt_wait: int,
    txt_attempts: int,
    txt_interval: int
):
    """Background task to run the pipeline with progress updates."""
    try:
        # Initialize status
        pipeline_status[run_id] = {
            "status": "running",
            "stage": "initializing",
            "started_at": datetime.now(timezone.utc).isoformat(),
            "total_domains": len(domains),
            "progress": 0
        }
        
        # Send WebSocket update
        await manager.send_progress(run_id, {
            "type": "status",
            "status": "running",
            "stage": "initializing",
            "message": f"Pipeline started for {len(domains)} domains",
            "progress": 0
        })
        
        # Create temporary CSV file with domains
        temp_csv = Path(f"data/temp_input_{run_id}.csv")
        temp_csv.parent.mkdir(parents=True, exist_ok=True)
        
        with open(temp_csv, 'w') as f:
            for i, domain in enumerate(domains, 1):
                f.write(f"{i},{domain},,\n")
        
        # Run pipeline
        pipeline = CompleteDomainPipeline(run_id)
        
        # Stage 1: API Lookup
        pipeline_status[run_id]["stage"] = "stage1_api"
        await manager.send_progress(run_id, {
            "type": "status",
            "stage": "stage1_api",
            "message": "Stage 1: RDAP/WHOIS API lookup",
            "progress": 10
        })
        
        await pipeline.run_complete_pipeline(
            input_csv=str(temp_csv),
            enable_txt_verification=enable_txt,
            txt_wait_time=txt_wait,
            txt_max_attempts=txt_attempts,
            txt_poll_interval=txt_interval
        )
        
        # Clean up temp file
        temp_csv.unlink(missing_ok=True)
        
        # Update status
        pipeline_status[run_id] = {
            "status": "completed",
            "stage": "finished",
            "finished_at": datetime.now(timezone.utc).isoformat(),
            "results_available": True,
            "progress": 100
        }
        
        # Send completion WebSocket message
        await manager.send_progress(run_id, {
            "type": "status",
            "status": "completed",
            "stage": "finished",
            "message": "Pipeline completed successfully",
            "progress": 100
        })
        
        # Send results to external APIs
        csv_path = Path(f"data/run_{run_id}/results/all_results_{run_id}.csv")
        if csv_path.exists():
            await send_to_external_apis(run_id, csv_path)
        
    except Exception as e:
        error_msg = str(e)
        pipeline_status[run_id] = {
            "status": "failed",
            "error": error_msg,
            "finished_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Send error WebSocket message
        await manager.send_progress(run_id, {
            "type": "error",
            "status": "failed",
            "message": f"Pipeline failed: {error_msg}",
            "error": error_msg
        })


@router.post("/run", response_model=PipelineRunResponse)
async def run_pipeline(request: PipelineRunRequest, background_tasks: BackgroundTasks):
    """
    Run the complete domain verification pipeline.
    
    This endpoint triggers a background task that runs:
    - Stage 1: RDAP/WHOIS API lookup
    - Stage 2: Playwright scraping (who.is + sidn.nl for .nl domains)
    - Stage 3: TXT verification setup
    - Stage 4: TXT verification execution (if enabled)
    
    Args:
        request: Pipeline run configuration
        background_tasks: FastAPI background tasks
        
    Returns:
        Pipeline run response with run_id
    """
    if not request.domains:
        raise HTTPException(status_code=400, detail="No domains provided")
    
    # Generate run ID
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Start background task
    background_tasks.add_task(
        run_pipeline_task,
        run_id,
        request.domains,
        request.enable_txt_verification,
        request.txt_wait_time,
        request.txt_max_attempts,
        request.txt_poll_interval
    )
    
    return PipelineRunResponse(
        run_id=run_id,
        status="started",
        message=f"Pipeline started for {len(request.domains)} domains",
        domains_count=len(request.domains),
        csv_download_url=f"/api/pipeline/{run_id}/csv",
        report_url=f"/api/pipeline/{run_id}/report",
        websocket_url=f"/api/pipeline/ws/{run_id}"
    )


@router.post("/upload", response_model=PipelineRunResponse)
async def upload_and_run_pipeline(
    file: UploadFile = File(...),
    enable_txt_verification: bool = False,
    background_tasks: BackgroundTasks = None
):
    """
    Upload a CSV/Excel file and run the pipeline.
    
    The file should contain a column with domain names.
    Supported formats: CSV, XLSX, XLS
    
    Args:
        file: Uploaded CSV or Excel file
        enable_txt_verification: Enable TXT verification (Stage 4)
        background_tasks: FastAPI background tasks
        
    Returns:
        Pipeline run response with run_id
    """
    suffix = Path(file.filename).suffix.lower()
    
    # Validate file type
    if suffix not in ('.csv', '.xlsx', '.xls'):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Only CSV and XLSX files are supported."
        )
    if suffix == '.xls':
        raise HTTPException(
            status_code=400,
            detail="Legacy .xls is not supported. Please upload .xlsx or .csv."
        )
    
    # Ensure background tasks available (defensive for direct invocation)
    if background_tasks is None:
        background_tasks = BackgroundTasks()
    
    # Generate run ID
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    try:
        # Save uploaded file temporarily
        temp_file = Path(f"data/temp_upload_{run_id}{suffix}")
        temp_file.parent.mkdir(parents=True, exist_ok=True)
        
        content = await file.read()
        with open(temp_file, 'wb') as f:
            f.write(content)
        
        # Convert XLSX to CSV if needed
        csv_source = temp_file
        temp_converted = None
        if suffix == '.xlsx':
            try:
                from openpyxl import load_workbook
            except ImportError:
                temp_file.unlink(missing_ok=True)
                raise HTTPException(status_code=500, detail="openpyxl is required to read .xlsx files")
            
            try:
                wb = load_workbook(temp_file, read_only=True, data_only=True)
                ws = wb.active
                
                temp_converted = Path(f"data/temp_upload_{run_id}.csv")
                with open(temp_converted, 'w', encoding='utf-8', newline='') as out_f:
                    writer = csv.writer(out_f)
                    for row in ws.iter_rows(values_only=True):
                        writer.writerow([
                            '' if cell is None else str(cell).strip() for cell in row
                        ])
                csv_source = temp_converted
            except Exception as e:
                temp_file.unlink(missing_ok=True)
                if temp_converted:
                    temp_converted.unlink(missing_ok=True)
                raise HTTPException(status_code=400, detail=f"Failed to convert XLSX: {str(e)}")
        
        # Parse domains from CSV source
        domains = []
        try:
            with open(csv_source, 'r', encoding='utf-8', errors='ignore', newline='') as f:
                reader = csv.reader(f)
                for row in reader:
                    if not row:
                        continue
                    domain = None
                    # Prefer second column if present, else first non-empty cell
                    if len(row) > 1 and row[1].strip():
                        domain = row[1].strip()
                    elif row[0].strip():
                        domain = row[0].strip()
                    
                    if domain and domain.lower() not in ['domain', 'domain_name']:
                        domains.append(domain)
        except Exception as e:
            temp_file.unlink(missing_ok=True)
            if temp_converted:
                temp_converted.unlink(missing_ok=True)
            raise HTTPException(status_code=400, detail=f"Failed to parse CSV: {str(e)}")
        
        if not domains:
            temp_file.unlink(missing_ok=True)
            raise HTTPException(status_code=400, detail="No valid domains found in file")
        
        # Start background task
        background_tasks.add_task(
            run_pipeline_task,
            run_id,
            domains,
            enable_txt_verification,
            30,  # txt_wait_time
            10,  # txt_max_attempts
            30   # txt_poll_interval
        )
        
        # Clean up temp upload files
        temp_file.unlink(missing_ok=True)
        if temp_converted:
            temp_converted.unlink(missing_ok=True)
        
        return PipelineRunResponse(
            run_id=run_id,
            status="started",
            message=f"Pipeline started for {len(domains)} domains from {file.filename}",
            domains_count=len(domains),
            csv_download_url=f"/api/pipeline/{run_id}/csv",
            report_url=f"/api/pipeline/{run_id}/report",
            websocket_url=f"/api/pipeline/ws/{run_id}"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


@router.websocket("/ws/{run_id}")
async def websocket_progress(websocket: WebSocket, run_id: str):
    """
    WebSocket endpoint for real-time pipeline progress updates.
    
    Clients can connect to this endpoint to receive live updates
    about the pipeline execution progress.
    
    Args:
        websocket: WebSocket connection
        run_id: Pipeline run ID
    """
    await manager.connect(run_id, websocket)
    try:
        # Send initial status
        if run_id in pipeline_status:
            await websocket.send_json({
                "type": "connected",
                "run_id": run_id,
                "status": pipeline_status[run_id].get("status", "unknown")
            })
        
        # Keep connection alive and listen for messages
        while True:
            data = await websocket.receive_text()
            # Echo back or handle commands if needed
            if data == "ping":
                await websocket.send_json({"type": "pong"})
                
    except WebSocketDisconnect:
        manager.disconnect(run_id)


@router.get("/status/{run_id}", response_model=PipelineStatusResponse)
async def get_pipeline_status(run_id: str):
    """
    Get the status of a pipeline run.
    
    Args:
        run_id: Pipeline run ID
        
    Returns:
        Pipeline status information
    """
    if run_id not in pipeline_status:
        # Check if run directory exists
        run_dir = Path(f"data/run_{run_id}")
        if run_dir.exists():
            return PipelineStatusResponse(
                run_id=run_id,
                status="completed",
                results_available=True,
                csv_url=f"/api/pipeline/{run_id}/csv",
                report_url=f"/api/pipeline/{run_id}/report"
            )
        else:
            raise HTTPException(status_code=404, detail="Run ID not found")
    
    status_info = pipeline_status[run_id]
    
    return PipelineStatusResponse(
        run_id=run_id,
        status=status_info["status"],
        stage=status_info.get("stage"),
        progress=status_info.get("progress"),
        results_available=status_info.get("results_available", False),
        csv_url=f"/api/pipeline/{run_id}/csv" if status_info.get("results_available") else None,
        report_url=f"/api/pipeline/{run_id}/report" if status_info.get("results_available") else None,
        error=status_info.get("error")
    )


@router.get("/{run_id}/csv")
async def download_csv(run_id: str):
    """
    Download CSV results for a pipeline run.
    
    Args:
        run_id: Pipeline run ID
        
    Returns:
        CSV file download
    """
    csv_file = Path(f"data/run_{run_id}/results/all_results_{run_id}.csv")
    
    if not csv_file.exists():
        raise HTTPException(status_code=404, detail="CSV file not found")
    
    return FileResponse(
        path=str(csv_file),
        media_type="text/csv",
        filename=f"domain_results_{run_id}.csv"
    )


@router.get("/{run_id}/report")
async def download_report(run_id: str):
    """
    Download text report for a pipeline run.
    
    Args:
        run_id: Pipeline run ID
        
    Returns:
        Text report file download
    """
    report_file = Path(f"data/run_{run_id}/results/FINAL_REPORT.txt")
    
    if not report_file.exists():
        raise HTTPException(status_code=404, detail="Report file not found")
    
    return FileResponse(
        path=str(report_file),
        media_type="text/plain",
        filename=f"report_{run_id}.txt"
    )


@router.get("/{run_id}/report/json")
async def download_report_json(run_id: str):
    """
    Download JSON report for a pipeline run.
    
    Args:
        run_id: Pipeline run ID
        
    Returns:
        JSON report file download
    """
    report_file = Path(f"data/run_{run_id}/results/FINAL_REPORT.json")
    
    if not report_file.exists():
        raise HTTPException(status_code=404, detail="JSON report file not found")
    
    return FileResponse(
        path=str(report_file),
        media_type="application/json",
        filename=f"report_{run_id}.json"
    )


@router.get("/{run_id}/screenshots/{filename}")
async def download_screenshot(run_id: str, filename: str):
    """
    Download a screenshot from a pipeline run.
    
    Args:
        run_id: Pipeline run ID
        filename: Screenshot filename
        
    Returns:
        Screenshot image file
    """
    screenshot_file = Path(f"data/run_{run_id}/screenshots/{filename}")
    
    if not screenshot_file.exists():
        raise HTTPException(status_code=404, detail="Screenshot not found")
    
    return FileResponse(
        path=str(screenshot_file),
        media_type="image/png"
    )


@router.post("/config/external-apis")
async def configure_external_apis(config: ExternalAPIConfig):
    """
    Configure external API endpoints for result delivery.
    
    This allows dynamic configuration of where pipeline results
    should be sent after completion.
    
    Args:
        config: External API configuration
        
    Returns:
        Success message
    """
    global external_api_config
    
    # Update configuration
    if config.momen_api_url:
        external_api_config.momen_api_url = config.momen_api_url
    if config.frontend_api_url:
        external_api_config.frontend_api_url = config.frontend_api_url
    if config.momen_api_key:
        external_api_config.momen_api_key = config.momen_api_key
    if config.frontend_api_key:
        external_api_config.frontend_api_key = config.frontend_api_key
    
    return JSONResponse(
        content={
            "message": "External API configuration updated",
            "momen_configured": bool(external_api_config.momen_api_url),
            "frontend_configured": bool(external_api_config.frontend_api_url)
        }
    )


@router.get("/config/external-apis")
async def get_external_api_config():
    """
    Get current external API configuration (without sensitive keys).
    
    Returns:
        Current configuration status
    """
    return JSONResponse(
        content={
            "momen_api_url": external_api_config.momen_api_url,
            "frontend_api_url": external_api_config.frontend_api_url,
            "momen_configured": bool(external_api_config.momen_api_url and external_api_config.momen_api_key),
            "frontend_configured": bool(external_api_config.frontend_api_url and external_api_config.frontend_api_key)
        }
    )
